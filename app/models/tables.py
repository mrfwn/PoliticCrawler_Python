from app import db
from openpyxl import load_workbook
from datetime import date , datetime , time ,timedelta
import json,requests
from sqlalchemy.orm import load_only
import time
from operator import itemgetter

class DeputadoF(db.Model):
    __tablename__ = "deputadof"

    #id = db.Column(db.Integer, primary_key = True)
    id_deputadof = db.Column(db.Integer,primary_key = True )
    #gastos = db.relationship('Gastos',backref='deputadof',lazy='dynamic')
    uri = db.Column(db.String)
    nome_civil = db.Column(db.String)
    nome_eleitoral = db.Column(db.String)
    sigla_partido = db.Column(db.String)
    uri_partido = db.Column(db.String)
    uf = db.Column(db.String)
    id_legislatura = db.Column(db.Integer)
    data_inicio_mandato = db.Column(db.String)
    url_foto = db.Column(db.String)
    end_predio = db.Column(db.String)
    end_sala = db.Column(db.String)
    end_andar = db.Column(db.String)
    end_telefone = db.Column(db.String)
    end_email = db.Column(db.String)
    situacao =  db.Column(db.String)
    condicao_eleitoral =  db.Column(db.String)
    sexo =  db.Column(db.String)
    data_nascimento = db.Column(db.String)
    uf_nascimento =  db.Column(db.String)
    municipio_nascimento =  db.Column(db.String)
    escolaridade =  db.Column(db.String)
    url_deputados_pe = "https://dadosabertos.camara.leg.br/api/v2/deputados?siglaUf=PE&ordem=ASC&ordenarPor=nome&pagina=1&itens=100"
    url_deputado = "https://dadosabertos.camara.leg.br/api/v2/deputados/"
    ocorrenciaQtd = {
        "bdpe":{
            "A":0, "B":0, "C":0, "X":0
        },
        "netv1":{
            "A":0, "B":0, "C":0, "X":0
        },
        "ge":{
            "A":0, "B":0, "C":0, "X":0
        },
        "netv2":{
            "A":0, "B":0, "C":0, "X":0
        }
    }

    ocorrenciaPorMes = {
        "bdpe":{
            "Janeiro":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Fevereiro":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Março":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Abril":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Maio":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Junho":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Julho":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Agosto":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Setembro":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Outubro":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Novembro":{
                "A":0, "B":0, "C":0, "X":0
            },
            "Dezembro":{
                "A":0, "B":0, "C":0, "X":0
            }
        }
    }
    
    def __init__(self,id_deputadof,uri,nome_civil,nome_eleitoral,sigla_partido,uri_partido,uf,
    id_legislatura,data_inicio_mandato,url_foto,end_predio,end_sala,end_andar,
    end_telefone,end_email,situacao,condicao_eleitoral,sexo,data_nascimento,uf_nascimento,municipio_nascimento,escolaridade):
        self.id_deputadof = id_deputadof
        self.uri = uri
        self.nome_civil = nome_civil
        self.nome_eleitoral = nome_eleitoral
        self.sigla_partido = sigla_partido
        self.uri_partido = uri_partido
        self.uf = uf
        self.id_legislatura = id_legislatura
        self.data_inicio_mandato = data_inicio_mandato
        self.url_foto = url_foto
        self.end_predio = end_predio
        self.end_sala = end_sala
        self.end_andar = end_andar
        self.end_telefone = end_telefone
        self.end_email = end_email
        self.situacao = situacao
        self.condicao_eleitoral = condicao_eleitoral
        self.sexo = sexo
        self.data_nascimento = data_nascimento
        self.uf_nascimento = uf_nascimento
        self.municipio_nascimento = municipio_nascimento
        self.escolaridade = escolaridade
        

    def __repr__(self):
        return "%r" % self.id_deputadof

    def getDeputados(self):
        count_dep = 0
        list_id_deputados = json.loads(requests.get(self.url_deputados_pe).content)
        while(count_dep < len(list_id_deputados['dados'])):
            url = self.url_deputado + str(list_id_deputados['dados'][count_dep]['id'])
            deputado = json.loads(requests.get(url).content)
            result = DeputadoF.query.filter_by(id_deputadof = deputado['dados']['id']).first()
            if result is None:
                try:
                    register = DeputadoF( 
                    deputado['dados']['id'],
                    deputado['dados']['uri'],
                    deputado['dados']['nomeCivil'],
                    deputado['dados']['ultimoStatus']['nomeEleitoral'],
                    deputado['dados']['ultimoStatus']['siglaPartido'], 
                    deputado['dados']['ultimoStatus']['uriPartido'],
                    deputado['dados']['ultimoStatus']['siglaUf'],
                    deputado['dados']['ultimoStatus']['idLegislatura'],
                    deputado['dados']['ultimoStatus']['data'],
                    deputado['dados']['ultimoStatus']['urlFoto'],
                    '','','','','',
                    deputado['dados']['ultimoStatus']['situacao'],
                    deputado['dados']['ultimoStatus']['condicaoEleitoral'],
                    deputado['dados']['sexo'],
                    deputado['dados']['dataNascimento'],
                    deputado['dados']['ufNascimento'],
                    deputado['dados']['municipioNascimento'],
                    deputado['dados']['escolaridade']
                    )
                    db.session.add(register)
                    db.session.commit()
                except:
                    db.rollback()
                    log.error('Rolling back transaction in query-none block')   
            count_dep+=1
        return 0

    def getIdsDep():
        url_deputados_pe = "https://dadosabertos.camara.leg.br/api/v2/deputados?siglaUf=PE&ordem=ASC&ordenarPor=nome&pagina=1&itens=100"
        list_id_deputados = json.loads(requests.get(url_deputados_pe).content)
        ids = []
        count_id = 0
        while(count_id < len(list_id_deputados['dados'])):           
            ids.append(int(list_id_deputados['dados'][count_id]['id']))           
            count_id+=1
        return ids 

    def loadPlanilha():
        line = 1
        xlsxInstance = load_workbook(filename='GFO - Relatório Detalhado.xlsx', read_only=True).active
        while (xlsxInstance.cell(row=line+1,column=1).value != None): 
            result = Gfo.query.filter_by(ocorrencia = xlsxInstance.cell(row=line,column=1).value).first()       
            
            if result is None:
                try:  
                    registro = Gfo(xlsxInstance.cell(row=line,column=1).value,xlsxInstance.cell(row=line,column=2).value)
                    db.session.add(registro)
                    db.session.commit()
                except:
                    db.rollback()
                    log.error('Rolling back transaction in query-none block')   

            
            else:
                try:    
                    """
                    session.query().\
                    filter(Class.ID == record.ID).\
                    update(some expression)
                    session.commit() 
                    """
                except:
                    db.rollback()
                    log.error('Rolling back transaction')
                    db.session.add(registro)
                    db.session.commit()
            line += 1


    def homeInfo(self):
        # Quantidades de Ocorrencias
        for x,a in self.ocorrenciaQtd.items():
            for y,b in self.ocorrenciaQtd[x].items():                
                self.ocorrenciaQtd[x][y] = Gfo.query.filter(Gfo.pgm == self.programAlias[x],Gfo.nivel_falha == y ).count()
        
        # Quantidade de Ocorrencias por mês
        for x,a in self.ocorrenciaPorMes.items():
            for y,b in self.ocorrenciaPorMes[x].items():
                for z,c in self.ocorrenciaPorMes[x][y].items():
                    self.ocorrenciaPorMes[x][y][z] = Gfo.query.filter(Gfo.pgm == self.programAlias[x] , Gfo.mes == y, Gfo.nivel_falha == z ).count()
       
        # Quantidade de Ocorrencias por tipo
        for x,a in self.ocorrenciaTipo.items():
            for y,b in self.ocorrenciaTipo[x].items():
                self.ocorrenciaTipo[x][y] = Gfo.query.filter(Gfo.abrangencia == x,Gfo.nivel_falha == y ).count()
        
        # Mtbf dos programas
        for x,a in self.mtbfValue.items():
            self.mtbfValue[x] = self.getMtbf(self,x)

        

class Gastos(db.Model):
    __tablename__ = "gastos"

    id = db.Column(db.Integer , primary_key=True)
    deputadof_id = db.Column(db.Integer, db.ForeignKey('deputadof.id_deputadof'))
    deputado = db.relationship('DeputadoF' , foreign_keys = deputadof_id)
    ano = db.Column(db.Integer)
    mes = db.Column(db.Integer)
    tipo_despesa = db.Column(db.String)
    id_documento = db.Column(db.Integer)
    tipo_documento = db.Column(db.String)
    id_tipo_documento = db.Column(db.Integer)
    data_documento = db.Column(db.String)
    numero_documento = db.Column(db.Integer)
    valor_documento = db.Column(db.Integer)
    nome_fornecedor = db.Column(db.String)     
    cnpj_cpf_fornecedor = db.Column(db.String)
    valor_liquido = db.Column(db.Integer)
    valor_glosa = db.Column(db.Integer)
    numero_ressarcimento = db.Column(db.Integer)
    id_lote = db.Column(db.Integer)
    parcela = db.Column(db.Integer)
    

    def __init__(self,deputadof_id,ano,mes,tipo_despesa,id_documento,tipo_documento,id_tipo_documento,
    data_documento,numero_documento,valor_documento,nome_fornecedor,cnpj_cpf_fornecedor,valor_liquido,
    valor_glosa,numero_ressarcimento,id_lote,parcela):
        self.deputadof_id = deputadof_id
        self.ano = ano
        self.mes = mes
        self.tipo_despesa = tipo_despesa
        self.id_documento = id_documento
        self.tipo_documento = tipo_documento
        self.id_tipo_documento = id_tipo_documento
        self.data_documento = data_documento
        self.numero_documento = numero_documento
        self.valor_documento = valor_documento
        self.nome_fornecedor = nome_fornecedor
        self.cnpj_cpf_fornecedor = cnpj_cpf_fornecedor
        self.valor_liquido = valor_liquido
        self.valor_glosa = valor_glosa
        self.numero_ressarcimento = numero_ressarcimento
        self.id_lote = id_lote
        self.parcela = parcela

    def __repr__(self):
        return "%s,%s,%r" % (self.deputado.nome_civil, self.tipo_despesa, self.valor_documento) 
 
    def getGastos(self):
        ids = DeputadoF.getIdsDep()
        count_dep = 0
        while(count_dep < len(ids)):
            page=1
            url = 'https://dadosabertos.camara.leg.br/api/v2/deputados/' + str(ids[count_dep]) + '/despesas?ano=2018&ordem=ASC&ordenarPor=ano&pagina=' + str(page) +'&itens=100'
            result = json.loads(requests.get(url).content)
            while(result['dados']!=[]):
                itens = 0
                while(itens<len(result['dados'])):
                    gastos = Gastos.query.filter_by(
                    deputadof_id = str(ids[count_dep]),
                    ano = result['dados'][itens]['ano'],
                    mes = result['dados'][itens]['mes'],
                    tipo_despesa = result['dados'][itens]['tipoDespesa'],
                    id_documento = result['dados'][itens]['idDocumento'],
                    tipo_documento = result['dados'][itens]['tipoDocumento'],
                    id_tipo_documento = result['dados'][itens]['idTipoDocumento'],
                    data_documento = result['dados'][itens]['dataDocumento'],
                    numero_documento = result['dados'][itens]['numDocumento'],
                    valor_documento = abs(result['dados'][itens]['valorDocumento']),
                    nome_fornecedor = result['dados'][itens]['nomeFornecedor'],
                    cnpj_cpf_fornecedor = result['dados'][itens]['cnpjCpfFornecedor'],
                    valor_liquido = result['dados'][itens]['valorLiquido'],
                    valor_glosa = result['dados'][itens]['valorGlosa'],
                    numero_ressarcimento = result['dados'][itens]['numRessarcimento'],
                    id_lote = result['dados'][itens]['idLote'],
                    parcela = result['dados'][itens]['parcela']
                    ).first()
                    if gastos is None:
                        try:
                            register = Gastos(
                            ids[count_dep],
                            result['dados'][itens]['ano'],
                            result['dados'][itens]['mes'],
                            result['dados'][itens]['tipoDespesa'],
                            result['dados'][itens]['idDocumento'],
                            result['dados'][itens]['tipoDocumento'],
                            result['dados'][itens]['idTipoDocumento'],
                            result['dados'][itens]['dataDocumento'],
                            result['dados'][itens]['numDocumento'],
                            abs(result['dados'][itens]['valorDocumento']),
                            result['dados'][itens]['nomeFornecedor'],
                            result['dados'][itens]['cnpjCpfFornecedor'],
                            abs(result['dados'][itens]['valorLiquido']),
                            result['dados'][itens]['valorGlosa'],
                            result['dados'][itens]['numRessarcimento'],
                            result['dados'][itens]['idLote'],
                            result['dados'][itens]['parcela']
                            )
                            db.session.add(register)
                            db.session.commit()
                        except:
                            db.rollback()
                            log.error('Rolling back transaction in query-none block') 
                    itens+=1
                page+=1 
                print(url)
                url = 'https://dadosabertos.camara.leg.br/api/v2/deputados/' + str(ids[count_dep]) + '/despesas?ano=2018&ordem=ASC&ordenarPor=ano&pagina=' + str(page) +'&itens=100'
                time.sleep(10)
                result = json.loads(requests.get(url).content)                
            count_dep+=1
        return 0
    
    def top10Aleimentacao():
        ids = DeputadoF.query.order_by(DeputadoF.id_deputadof).all()
        tops = []
        count_ids = 0
        while(count_ids < len(ids)):
            var = Gastos.query.filter(Gastos.deputado == ids[count_ids]).filter(Gastos.tipo_despesa == 'FORNECIMENTO DE ALIMENTAÇÃO DO PARLAMENTAR').all()  
            count_reg = 0
            candidate = {
            'Nome': '',
            'Gasto': '',
            'Total': 0
            }
            while(count_reg < len(var)):
                candidate['Nome']=var[count_reg].deputado.nome_civil
                candidate['Gasto']=var[count_reg].tipo_despesa
                candidate['Total'] += var[count_reg].valor_documento 
                count_reg+=1
            tops.append(candidate)
            count_ids+=1
        return tops
    
    def top10Combustivel():
        ids = DeputadoF.query.order_by(DeputadoF.id_deputadof).all()
        tops = []
        count_ids = 0
        while(count_ids < len(ids)):
            var = Gastos.query.filter(Gastos.deputado == ids[count_ids]).filter(Gastos.tipo_despesa == 'COMBUSTÍVEIS E LUBRIFICANTES.').all()  
            count_reg = 0
            candidate = {
            'Nome': '',
            'Gasto': '',
            'Total': 0
            }
            while(count_reg < len(var)):
                candidate['Nome']=var[count_reg].deputado.nome_civil
                candidate['Gasto']=var[count_reg].tipo_despesa
                candidate['Total'] += var[count_reg].valor_documento 
                count_reg+=1
            tops.append(candidate)
            count_ids+=1
        return tops
     
"""
class User(db.Model):
    __tablename__ = "users"

    id = db.Column(db.Integer , primary_key=True)
    username = db.Column(db.String, unique=True)
    password = db.Column(db.String)
    name = db.Column(db.String)
    email = db.Column(db.String)

    def __init__(self, username,password,name,email):
        self.username=username
        self.password = password
        self.name = name
        self.email = email

    def __repr__(self):
        return "<User %r>" % self.username        
"""




"""
class Follow(db.Model):
    __tablename__ = "follow"

    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey('users.id'))
    follower_id = db.Column(db.Integer,db.ForeignKey('users.id'))

    user = db.relationship('User', foreign_keys=user_id)
    follower = db.relationship('User', foreign_keys=follower_id)

"""