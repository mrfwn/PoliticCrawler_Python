import json,requests
url = "https://dadosabertos.camara.leg.br/api/v2/deputados?siglaUf=PE&itens=600&ordem=ASC&ordenarPor=nome"

"""
session = requests.session()
session.headers = {'User-Agent': 'Mozilla/5.0'}
"""
resp = requests.get(url)
value = json.loads(resp.content)
print(value['dados'][0]['id'])
#value = json.loads(resp.content)
#print(value['dados'])

"""
import urllib.request
from openpyxl import load_workbook
urllib.request.urlretrieve("http://www.recife.pe.leg.br/portal-da-transparencia/verba-indenizatoria-1/outubro-1", "file.xlsx")

wb = load_workbook("file.xlsx")
ws = wb["WANDERSON SOBRAL"]

print(ws['A1'].value)

"""